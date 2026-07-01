from __future__ import annotations

"""
DemoSQLQueryTool
================

這是將原本 HTML + JavaScript 版本改寫成的 Python 桌面程式。

主要功能：
1. 讀取 MySQL Dump 格式的 demo.sql。
2. 顯示資料表與欄位。
3. 支援兩組 WHERE 條件。
4. 支援 AND、OR、NOT。
5. 支援 LIKE、BETWEEN、IS NULL、ORDER BY。
6. 顯示對應的 SQL 預覽。
7. 將目前查詢結果匯出成 Excel。
8. 可使用 PyInstaller 包裝成 Windows .exe。

重要說明：
- 這個程式不會連接 MySQL Server。
- 它會解析 demo.sql 中的 CREATE TABLE 與 INSERT INTO。
- 查詢由 Python 模擬執行，但畫面會顯示對應的 SQL。
"""

import math
import os
import re
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from functools import cmp_to_key
from pathlib import Path
from typing import Any

import tkinter as tk
from tkinter import filedialog, messagebox, ttk


# ============================================================
# 1. 程式共用常數
# ============================================================

PREFERRED_TABLE_ORDER = [
    "categories",
    "customers",
    "employees",
    "orderdetails",
    "orders",
    "products",
    "shippers",
    "suppliers",
]

NOT_OPTIONS = [
    "一般條件",
    "NOT（反轉條件）",
]

LOGICAL_OPTIONS = [
    "AND（兩個都要符合）",
    "OR（任一符合即可）",
]

OPERATOR_LABELS = {
    "like": "LIKE（自行輸入 % 或 _）",
    "eq": "等於（=）",
    "neq": "不等於（<>）",
    "gt": "大於（>）",
    "gte": "大於等於（>=）",
    "lt": "小於（<）",
    "lte": "小於等於（<=）",
    "between": "介於（BETWEEN）",
    "isnull": "是 NULL",
    "notnull": "不是 NULL",
}

LABEL_TO_OPERATOR = {
    label: code
    for code, label in OPERATOR_LABELS.items()
}

SORT_DIRECTIONS = [
    "ASC（正序）",
    "DESC（反序）",
]

SORT_MODES = [
    "依原欄位型態",
    "轉成數字排序",
]

SECOND_COLUMN_NONE = "不使用第二條件"
SORT_COLUMN_NONE = "不排序"

# 條件未啟用時使用的特殊物件。
# 它與 SQL 的 NULL / UNKNOWN 不同。
INACTIVE_CONDITION = object()

# 判斷未加引號數值的正規表示式。
NUMBER_PATTERN = re.compile(
    r"^-?(?:\d+\.?\d*|\.\d+)(?:e[+-]?\d+)?$",
    re.IGNORECASE,
)


# ============================================================
# 2. 資料表資料結構
# ============================================================

@dataclass
class TableData:
    """
    保存一張資料表的解析結果。

    columns：
    欄位順序。

    types：
    欄位名稱對應到 MySQL 資料型態。

    rows：
    每一筆資料都是一個 list。
    """

    columns: list[str]
    types: dict[str, str]
    rows: list[list[Any]]


# ============================================================
# 3. 尋找程式旁邊或 EXE 內的 demo.sql
# ============================================================

def application_directory() -> Path:
    """
    回傳程式所在資料夾。

    一般執行 app.py：
    回傳 app.py 所在資料夾。

    PyInstaller EXE：
    回傳 EXE 所在資料夾。
    """

    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent


def bundled_directory() -> Path:
    """
    PyInstaller --onefile 會將內含資源暫時解壓縮到 _MEIPASS。

    若目前不是 PyInstaller EXE，
    就回傳 app.py 所在資料夾。
    """

    temp_directory = getattr(sys, "_MEIPASS", None)

    if temp_directory:
        return Path(temp_directory)

    return Path(__file__).resolve().parent


def find_default_sql_file() -> Path | None:
    """
    尋找預設 demo.sql。

    尋找順序：
    1. EXE 或 app.py 旁邊的 demo.sql。
    2. PyInstaller EXE 內嵌的 demo.sql。

    外部檔案優先，方便使用者直接更換 demo.sql。
    """

    external_sql = application_directory() / "demo.sql"

    if external_sql.exists():
        return external_sql

    embedded_sql = bundled_directory() / "demo.sql"

    if embedded_sql.exists():
        return embedded_sql

    return None


# ============================================================
# 4. 解析 MySQL 字串跳脫字元
# ============================================================

def decode_mysql_escape(character: str) -> str:
    """
    將 MySQL Dump 字串中的跳脫字元轉回真正字元。

    例如：
    \\n 轉為換行
    \\t 轉為 Tab
    \\' 轉為單引號
    """

    escapes = {
        "0": "\0",
        "b": "\b",
        "n": "\n",
        "r": "\r",
        "t": "\t",
        "Z": "\x1a",
        "\\": "\\",
        "'": "'",
        '"': '"',
    }

    return escapes.get(character, character)


# ============================================================
# 5. 解析 INSERT INTO ... VALUES (...)
# ============================================================

def parse_value_tuples(values_text: str) -> list[list[Any]]:
    """
    將 SQL VALUES 內容轉成 Python 二維 list。

    SQL：
    (1,'Chais',18.00),(2,'Chang',19.00)

    轉換後：
    [
        [1, "Chais", 18.0],
        [2, "Chang", 19.0]
    ]
    """

    rows: list[list[Any]] = []
    index = 0
    text_length = len(values_text)

    def skip_whitespace() -> None:
        """跳過空格、換行與 Tab。"""

        nonlocal index

        while (
            index < text_length
            and values_text[index].isspace()
        ):
            index += 1

    while index < text_length:
        skip_whitespace()

        # 不同 Row 之間通常由逗號分隔。
        if (
            index < text_length
            and values_text[index] == ","
        ):
            index += 1
            continue

        # 每一筆 Row 必須由左括號開始。
        if (
            index >= text_length
            or values_text[index] != "("
        ):
            index += 1
            continue

        # 略過左括號。
        index += 1

        row: list[Any] = []

        while index < text_length:
            skip_whitespace()

            value: Any

            # 單引號開頭代表 SQL 文字。
            if values_text[index] == "'":
                index += 1
                characters: list[str] = []

                while index < text_length:
                    current = values_text[index]

                    # MySQL 反斜線跳脫。
                    if current == "\\":
                        index += 1

                        if index < text_length:
                            characters.append(
                                decode_mysql_escape(
                                    values_text[index]
                                )
                            )
                            index += 1

                        continue

                    # 單引號可能是字串結尾，
                    # 也可能是 SQL 的兩個單引號跳脫。
                    if current == "'":
                        if (
                            index + 1 < text_length
                            and values_text[index + 1] == "'"
                        ):
                            characters.append("'")
                            index += 2
                            continue

                        index += 1
                        break

                    characters.append(current)
                    index += 1

                value = "".join(characters)

            else:
                # 未加引號內容可能是數字或 NULL。
                start = index

                while (
                    index < text_length
                    and values_text[index] not in ",)"
                ):
                    index += 1

                token = values_text[start:index].strip()

                if token.upper() == "NULL":
                    value = None

                elif NUMBER_PATTERN.fullmatch(token):
                    # 整數保留為 int，小數保留為 float。
                    if (
                        "." not in token
                        and "e" not in token.lower()
                    ):
                        value = int(token)
                    else:
                        value = float(token)

                else:
                    value = token

            row.append(value)
            skip_whitespace()

            # 逗號表示後面還有下一個欄位。
            if (
                index < text_length
                and values_text[index] == ","
            ):
                index += 1
                continue

            # 右括號表示目前 Row 結束。
            if (
                index < text_length
                and values_text[index] == ")"
            ):
                index += 1
                rows.append(row)
                break

    return rows


# ============================================================
# 6. 解析整份 MySQL Dump
# ============================================================

def parse_sql_dump(sql_text: str) -> dict[str, TableData]:
    """
    解析 demo.sql 中的：

    CREATE TABLE
    INSERT INTO ... VALUES

    回傳：
    {
        "customers": TableData(...),
        "products": TableData(...)
    }
    """

    parsed: dict[str, TableData] = {}

    create_pattern = re.compile(
        r"CREATE TABLE\s+`([^`]+)`\s*"
        r"\(([\s\S]*?)\)\s*ENGINE=",
        re.IGNORECASE,
    )

    for create_match in create_pattern.finditer(sql_text):
        table_name = create_match.group(1)
        definition = create_match.group(2)

        columns: list[str] = []
        types: dict[str, str] = {}

        # 每一行欄位定義通常類似：
        # `CustomerID` varchar(10) NOT NULL
        column_pattern = re.compile(
            r"^\s*`([^`]+)`\s+"
            r"([a-zA-Z]+(?:\s*\([^)]*\))?)",
            re.IGNORECASE,
        )

        for line in definition.splitlines():
            column_match = column_pattern.match(line)

            # PRIMARY KEY、KEY 等行不符合欄位格式。
            if not column_match:
                continue

            column_name = column_match.group(1)

            data_type = re.sub(
                r"\s+",
                "",
                column_match.group(2),
            ).lower()

            columns.append(column_name)
            types[column_name] = data_type

        parsed[table_name] = TableData(
            columns=columns,
            types=types,
            rows=[],
        )

    insert_pattern = re.compile(
        r"INSERT INTO\s+`([^`]+)`\s+"
        r"VALUES\s*([\s\S]*?);",
        re.IGNORECASE,
    )

    for insert_match in insert_pattern.finditer(sql_text):
        table_name = insert_match.group(1)
        values_text = insert_match.group(2)
        rows = parse_value_tuples(values_text)

        if table_name not in parsed:
            parsed[table_name] = TableData(
                columns=[],
                types={},
                rows=[],
            )

        parsed[table_name].rows.extend(rows)

    return parsed


# ============================================================
# 7. SQL 型態與 SQL 文字處理
# ============================================================

def is_numeric_type(type_name: str) -> bool:
    """判斷 MySQL 欄位是否為常見數字型態。"""

    return bool(
        re.match(
            r"^(tinyint|smallint|mediumint|int|integer|"
            r"bigint|decimal|numeric|float|double|real)",
            type_name or "",
            re.IGNORECASE,
        )
    )


def quote_sql_text(value: Any) -> str:
    """
    將 Python 文字轉成 SQL 字串。

    O'Brien
    轉為：
    'O''Brien'
    """

    escaped = str(value).replace("'", "''")
    return f"'{escaped}'"


def sql_literal(value: Any, type_name: str) -> str:
    """
    根據欄位型態決定 SQL 值是否加單引號。

    數字欄位：
    50

    文字欄位：
    'Germany'
    """

    trimmed = str(value).strip()

    if (
        is_numeric_type(type_name)
        and trimmed
    ):
        try:
            Decimal(trimmed)
            return trimmed
        except InvalidOperation:
            pass

    return quote_sql_text(trimmed)


def condition_is_ready(
    column_name: str,
    operator: str,
    first_value: str,
    second_value: str,
) -> bool:
    """
    判斷條件是否已填寫完整。

    IS NULL / IS NOT NULL 不需要搜尋值。
    BETWEEN 需要兩個搜尋值。
    """

    if not column_name:
        return False

    if operator in {"isnull", "notnull"}:
        return True

    if not first_value.strip():
        return False

    if (
        operator == "between"
        and not second_value.strip()
    ):
        return False

    return True


def build_condition_sql(
    column_name: str,
    operator: str,
    first_value: str,
    second_value: str,
    table_data: TableData,
) -> str:
    """
    將一組 UI 條件轉成 SQL 條件文字。

    此函式不會加入 WHERE，
    因為主程式還要處理 AND / OR / NOT。
    """

    if not condition_is_ready(
        column_name,
        operator,
        first_value,
        second_value,
    ):
        return ""

    type_name = table_data.types.get(
        column_name,
        "",
    )

    if operator == "isnull":
        return f"{column_name} IS NULL"

    if operator == "notnull":
        return f"{column_name} IS NOT NULL"

    if operator == "eq":
        return (
            f"{column_name} = "
            f"{sql_literal(first_value, type_name)}"
        )

    if operator == "neq":
        return (
            f"{column_name} <> "
            f"{sql_literal(first_value, type_name)}"
        )

    if operator == "gt":
        return (
            f"{column_name} > "
            f"{sql_literal(first_value, type_name)}"
        )

    if operator == "gte":
        return (
            f"{column_name} >= "
            f"{sql_literal(first_value, type_name)}"
        )

    if operator == "lt":
        return (
            f"{column_name} < "
            f"{sql_literal(first_value, type_name)}"
        )

    if operator == "lte":
        return (
            f"{column_name} <= "
            f"{sql_literal(first_value, type_name)}"
        )

    if operator == "like":
        # % 與 _ 由使用者自行輸入。
        return (
            f"{column_name} LIKE "
            f"{quote_sql_text(first_value)}"
        )

    if operator == "between":
        return (
            f"{column_name} BETWEEN "
            f"{sql_literal(first_value, type_name)} "
            f"AND "
            f"{sql_literal(second_value, type_name)}"
        )

    # 未知代號預設使用等於。
    return (
        f"{column_name} = "
        f"{sql_literal(first_value, type_name)}"
    )


# ============================================================
# 8. Python 模擬 SQL 比較
# ============================================================

def comparable_values(
    left: Any,
    right: Any,
    type_name: str,
) -> tuple[Any, Any] | None:
    """
    將左右兩個值轉成可比較格式。

    數字型態：
    轉成 Decimal。

    文字型態：
    轉成 casefold()，達到不分英文大小寫。
    """

    if left is None:
        # SQL 中 NULL 與一般值比較會得到 UNKNOWN。
        return None

    if is_numeric_type(type_name):
        try:
            return (
                Decimal(str(left)),
                Decimal(str(right).strip()),
            )
        except InvalidOperation:
            pass

    return (
        str(left).casefold(),
        str(right).casefold(),
    )


def like_pattern_to_regex(pattern: str) -> re.Pattern[str]:
    """
    將 SQL LIKE 萬用字元轉成 Python 正規表示式。

    % → 任意數量字元
    _ → 任意一個字元
    """

    regex_parts = ["^"]

    for character in str(pattern):
        if character == "%":
            regex_parts.append(".*")
        elif character == "_":
            regex_parts.append(".")
        else:
            regex_parts.append(re.escape(character))

    regex_parts.append("$")

    return re.compile(
        "".join(regex_parts),
        re.IGNORECASE | re.DOTALL,
    )


def evaluate_condition(
    row: list[Any],
    table_data: TableData,
    column_name: str,
    operator: str,
    first_value: str,
    second_value: str,
) -> bool | None | object:
    """
    判斷一筆 Row 是否符合一組條件。

    回傳 True：
    條件成立。

    回傳 False：
    條件不成立。

    回傳 None：
    SQL UNKNOWN，通常是一般比較遇到 NULL。

    回傳 INACTIVE_CONDITION：
    條件未啟用或尚未填完整。
    """

    if not condition_is_ready(
        column_name,
        operator,
        first_value,
        second_value,
    ):
        return INACTIVE_CONDITION

    try:
        column_index = table_data.columns.index(
            column_name
        )
    except ValueError:
        return INACTIVE_CONDITION

    cell_value = row[column_index]
    type_name = table_data.types.get(
        column_name,
        "",
    )

    if operator == "isnull":
        return cell_value is None

    if operator == "notnull":
        return cell_value is not None

    # SQL 中 NULL 使用一般比較會得到 UNKNOWN。
    if cell_value is None:
        return None

    pair = comparable_values(
        cell_value,
        first_value,
        type_name,
    )

    if pair is None:
        return None

    left, right = pair

    if operator == "eq":
        return left == right

    if operator == "neq":
        return left != right

    if operator == "gt":
        return left > right

    if operator == "gte":
        return left >= right

    if operator == "lt":
        return left < right

    if operator == "lte":
        return left <= right

    if operator == "like":
        return bool(
            like_pattern_to_regex(first_value).match(
                str(cell_value)
            )
        )

    if operator == "between":
        lower_pair = comparable_values(
            cell_value,
            first_value,
            type_name,
        )

        upper_pair = comparable_values(
            cell_value,
            second_value,
            type_name,
        )

        if (
            lower_pair is None
            or upper_pair is None
        ):
            return None

        return (
            lower_pair[0] >= lower_pair[1]
            and upper_pair[0] <= upper_pair[1]
        )

    return left == right


def apply_sql_not(
    result: bool | None | object,
    use_not: bool,
) -> bool | None | object:
    """
    套用 SQL NOT。

    SQL 三值邏輯：
    NOT TRUE    → FALSE
    NOT FALSE   → TRUE
    NOT UNKNOWN → UNKNOWN

    未啟用的條件也不做反轉。
    """

    if not use_not:
        return result

    if (
        result is INACTIVE_CONDITION
        or result is None
    ):
        return result

    return not result


def combine_sql_results(
    first: bool | None,
    second: bool | None,
    logical_operator: str,
) -> bool | None:
    """
    依 SQL 三值邏輯合併 AND / OR。

    AND：
    只要有 FALSE 就是 FALSE；
    兩個 TRUE 才是 TRUE；
    其餘是 UNKNOWN。

    OR：
    只要有 TRUE 就是 TRUE；
    兩個 FALSE 才是 FALSE；
    其餘是 UNKNOWN。
    """

    if logical_operator == "OR":
        if first is True or second is True:
            return True

        if first is False and second is False:
            return False

        return None

    # 預設使用 AND。
    if first is False or second is False:
        return False

    if first is True and second is True:
        return True

    return None


# ============================================================
# 9. 排序
# ============================================================

def sort_rows(
    rows: list[list[Any]],
    table_data: TableData,
    column_name: str,
    direction: str,
    force_numeric: bool,
) -> list[list[Any]]:
    """
    模擬 SQL ORDER BY。

    NULL 固定放在最後。
    """

    if not column_name:
        return list(rows)

    try:
        column_index = table_data.columns.index(
            column_name
        )
    except ValueError:
        return list(rows)

    type_name = table_data.types.get(
        column_name,
        "",
    )

    direction_factor = (
        -1
        if direction == "DESC"
        else 1
    )

    def compare_rows(
        left_row: list[Any],
        right_row: list[Any],
    ) -> int:
        left = left_row[column_index]
        right = right_row[column_index]

        if left is None and right is None:
            return 0

        if left is None:
            return 1

        if right is None:
            return -1

        if (
            force_numeric
            or is_numeric_type(type_name)
        ):
            try:
                left_number = Decimal(str(left))
                right_number = Decimal(str(right))

                if left_number < right_number:
                    return -1 * direction_factor

                if left_number > right_number:
                    return 1 * direction_factor

                return 0
            except InvalidOperation:
                pass

        left_text = str(left).casefold()
        right_text = str(right).casefold()

        if left_text < right_text:
            return -1 * direction_factor

        if left_text > right_text:
            return 1 * direction_factor

        return 0

    return sorted(
        rows,
        key=cmp_to_key(compare_rows),
    )


# ============================================================
# 10. Excel 匯出
# ============================================================

def safe_excel_sheet_name(name: str) -> str:
    """
    Excel 工作表名稱最多 31 字，
    且不能包含反斜線、斜線、問號、星號、中括號與冒號。
    """

    cleaned = re.sub(
        r'[\\/?*\[\]:]',
        "_",
        str(name),
    )[:31]

    return cleaned or "QueryResult"


def export_result_to_excel(
    output_path: Path,
    table_name: str,
    columns: list[str],
    rows: list[list[Any]],
    sql_text: str,
) -> None:
    """
    產生真正的 .xlsx。

    Excel 內容：
    第 1 列起：SELECT / FROM / WHERE / ORDER BY
    空一列
    欄位標題
    查詢結果
    """

    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "缺少 openpyxl，請先執行："
            "py -m pip install openpyxl"
        ) from error

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = safe_excel_sheet_name(
        table_name
    )

    sql_lines = (
        sql_text.splitlines()
        if sql_text.strip()
        else ["SELECT *", f"FROM demo.{table_name};"]
    )

    last_column = max(len(columns), 1)

    sql_fill = PatternFill(
        fill_type="solid",
        fgColor="111827",
    )

    sql_font = Font(
        color="E5E7EB",
        bold=True,
    )

    # SQL 每一行放在 Excel 上方各一列。
    for row_number, sql_line in enumerate(
        sql_lines,
        start=1,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=sql_line,
        )

        if last_column > 1:
            worksheet.merge_cells(
                start_row=row_number,
                start_column=1,
                end_row=row_number,
                end_column=last_column,
            )

        cell = worksheet.cell(
            row=row_number,
            column=1,
        )

        cell.fill = sql_fill
        cell.font = sql_font
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )

        worksheet.row_dimensions[
            row_number
        ].height = 22

    # SQL 後空一列，再放欄位標題。
    header_row = len(sql_lines) + 2

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    thin_gray = Side(
        style="thin",
        color="D9E2F3",
    )

    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    for column_number, column_name in enumerate(
        columns,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_number,
            value=column_name,
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # 寫入目前查詢結果。
    for row_offset, source_row in enumerate(
        rows,
        start=1,
    ):
        excel_row = header_row + row_offset

        for column_number in range(
            1,
            len(columns) + 1,
        ):
            value = (
                source_row[column_number - 1]
                if column_number - 1 < len(source_row)
                else None
            )

            # 畫面中的 NULL 在 Excel 中也顯示成 NULL。
            excel_value = (
                "NULL"
                if value is None
                else value
            )

            cell = worksheet.cell(
                row=excel_row,
                column=column_number,
                value=excel_value,
            )

            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
            )

    # 自動篩選範圍。
    if columns:
        end_row = max(
            header_row,
            header_row + len(rows),
        )

        worksheet.auto_filter.ref = (
            f"A{header_row}:"
            f"{get_column_letter(len(columns))}"
            f"{end_row}"
        )

    # 捲動時固定欄位標題。
    worksheet.freeze_panes = (
        f"A{header_row + 1}"
    )

    # 估算欄寬，限制在 10 至 40。
    for index, column_name in enumerate(
        columns,
        start=1,
    ):
        maximum_length = len(str(column_name))

        for source_row in rows:
            if index - 1 >= len(source_row):
                continue

            value = source_row[index - 1]
            value_text = (
                "NULL"
                if value is None
                else str(value)
            )

            maximum_length = max(
                maximum_length,
                len(value_text),
            )

        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = min(
            max(maximum_length + 2, 10),
            40,
        )

    workbook.save(output_path)


# ============================================================
# 11. Python 桌面 UI
# ============================================================

class DemoSqlQueryApp:
    """
    主視窗類別。

    UI 使用 Python 內建 tkinter / ttk，
    因此不需要瀏覽器或 HTML。
    """

    def __init__(self, root: tk.Tk) -> None:
        self.root = root

        self.root.title(
            "demo.sql SQL 查詢工具"
        )

        self.root.geometry("1450x880")
        self.root.minsize(1050, 680)

        # 解析後的所有資料表。
        self.database: dict[str, TableData] = {}

        # 目前畫面顯示的查詢結果。
        self.current_result: list[list[Any]] = []

        # 最後一次真正執行的 SQL。
        self.last_executed_sql = ""

        # 目前載入的 .sql 路徑。
        self.current_sql_path: Path | None = None

        # 顯示文字 → 真正欄位名稱。
        self.column_display_map: dict[str, str] = {}

        self._create_variables()
        self._create_styles()
        self._create_widgets()
        self._bind_events()

        # 等視窗建立完成後，自動讀取 demo.sql。
        self.root.after(
            100,
            self.load_default_sql,
        )

    # --------------------------------------------------------
    # 11-1. tkinter 變數
    # --------------------------------------------------------

    def _create_variables(self) -> None:
        """建立所有會與 UI 元件連動的變數。"""

        self.table_var = tk.StringVar()

        self.column1_var = tk.StringVar()
        self.not1_var = tk.StringVar(
            value=NOT_OPTIONS[0]
        )
        self.operator1_var = tk.StringVar(
            value=OPERATOR_LABELS["like"]
        )
        self.value1_var = tk.StringVar()
        self.value2_var = tk.StringVar()

        self.logical_var = tk.StringVar(
            value=LOGICAL_OPTIONS[0]
        )

        self.column2_var = tk.StringVar(
            value=SECOND_COLUMN_NONE
        )
        self.not2_var = tk.StringVar(
            value=NOT_OPTIONS[0]
        )
        self.operator2_var = tk.StringVar(
            value=OPERATOR_LABELS["like"]
        )
        self.value3_var = tk.StringVar()
        self.value4_var = tk.StringVar()

        self.sort_column_var = tk.StringVar(
            value=SORT_COLUMN_NONE
        )
        self.sort_direction_var = tk.StringVar(
            value=SORT_DIRECTIONS[0]
        )
        self.sort_mode_var = tk.StringVar(
            value=SORT_MODES[0]
        )

        self.load_status_var = tk.StringVar(
            value="準備讀取資料……"
        )

        self.result_status_var = tk.StringVar()

    # --------------------------------------------------------
    # 11-2. ttk 外觀
    # --------------------------------------------------------

    def _create_styles(self) -> None:
        """設定基本 UI 外觀。"""

        style = ttk.Style()

        # Windows 上優先使用較現代的 vista 風格。
        available = style.theme_names()

        if "vista" in available:
            style.theme_use("vista")
        elif "clam" in available:
            style.theme_use("clam")

        style.configure(
            "Title.TLabel",
            font=("Microsoft JhengHei UI", 17, "bold"),
        )

        style.configure(
            "Description.TLabel",
            font=("Microsoft JhengHei UI", 10),
            foreground="#475569",
        )

        style.configure(
            "Primary.TButton",
            font=("Microsoft JhengHei UI", 10, "bold"),
        )

        style.configure(
            "Treeview",
            rowheight=27,
            font=("Microsoft JhengHei UI", 10),
        )

        style.configure(
            "Treeview.Heading",
            font=("Microsoft JhengHei UI", 10, "bold"),
        )

    # --------------------------------------------------------
    # 11-3. 建立 UI 元件
    # --------------------------------------------------------

    def _create_widgets(self) -> None:
        """建立整個視窗內容。"""

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        main = ttk.Frame(
            self.root,
            padding=14,
        )

        main.grid(
            row=0,
            column=0,
            sticky="nsew",
        )

        main.columnconfigure(0, weight=1)
        main.rowconfigure(6, weight=1)

        # -------------------------
        # 標題
        # -------------------------

        ttk.Label(
            main,
            text="demo.sql SQL 查詢工具",
            style="Title.TLabel",
        ).grid(
            row=0,
            column=0,
            sticky="w",
        )

        ttk.Label(
            main,
            text=(
                "支援兩組 WHERE 條件、AND／OR／NOT、"
                "LIKE、BETWEEN、ORDER BY 及 Excel 匯出。"
            ),
            style="Description.TLabel",
        ).grid(
            row=1,
            column=0,
            sticky="w",
            pady=(2, 10),
        )

        # -------------------------
        # 檔案與資料表
        # -------------------------

        source_frame = ttk.LabelFrame(
            main,
            text="資料來源",
            padding=10,
        )

        source_frame.grid(
            row=2,
            column=0,
            sticky="ew",
            pady=(0, 8),
        )

        source_frame.columnconfigure(1, weight=1)

        ttk.Label(
            source_frame,
            text="FROM：資料表",
        ).grid(
            row=0,
            column=0,
            sticky="w",
            padx=(0, 8),
        )

        self.table_combo = ttk.Combobox(
            source_frame,
            textvariable=self.table_var,
            state="disabled",
            width=28,
        )

        self.table_combo.grid(
            row=0,
            column=1,
            sticky="w",
        )

        ttk.Button(
            source_frame,
            text="選擇 SQL 檔",
            command=self.choose_sql_file,
        ).grid(
            row=0,
            column=2,
            padx=(12, 4),
        )

        ttk.Button(
            source_frame,
            text="重新讀取",
            command=self.reload_sql_file,
        ).grid(
            row=0,
            column=3,
            padx=4,
        )

        # -------------------------
        # WHERE 條件區
        # -------------------------

        conditions = ttk.Frame(main)

        conditions.grid(
            row=3,
            column=0,
            sticky="ew",
            pady=(0, 8),
        )

        conditions.columnconfigure(0, weight=1)
        conditions.columnconfigure(1, weight=0)
        conditions.columnconfigure(2, weight=1)

        # 條件 1
        condition1_frame = ttk.LabelFrame(
            conditions,
            text="WHERE 條件 1",
            padding=10,
        )

        condition1_frame.grid(
            row=0,
            column=0,
            sticky="nsew",
            padx=(0, 6),
        )

        for column in range(5):
            condition1_frame.columnconfigure(
                column,
                weight=1,
            )

        ttk.Label(
            condition1_frame,
            text="欄位",
        ).grid(
            row=0,
            column=0,
            sticky="w",
        )

        ttk.Label(
            condition1_frame,
            text="NOT",
        ).grid(
            row=0,
            column=1,
            sticky="w",
        )

        ttk.Label(
            condition1_frame,
            text="比較方式",
        ).grid(
            row=0,
            column=2,
            sticky="w",
        )

        ttk.Label(
            condition1_frame,
            text="搜尋值",
        ).grid(
            row=0,
            column=3,
            sticky="w",
        )

        ttk.Label(
            condition1_frame,
            text="BETWEEN 第二值",
        ).grid(
            row=0,
            column=4,
            sticky="w",
        )

        self.column1_combo = ttk.Combobox(
            condition1_frame,
            textvariable=self.column1_var,
            state="disabled",
            width=24,
        )

        self.column1_combo.grid(
            row=1,
            column=0,
            sticky="ew",
            padx=(0, 5),
        )

        self.not1_combo = ttk.Combobox(
            condition1_frame,
            textvariable=self.not1_var,
            values=NOT_OPTIONS,
            state="disabled",
            width=15,
        )

        self.not1_combo.grid(
            row=1,
            column=1,
            sticky="ew",
            padx=5,
        )

        self.operator1_combo = ttk.Combobox(
            condition1_frame,
            textvariable=self.operator1_var,
            values=list(OPERATOR_LABELS.values()),
            state="disabled",
            width=22,
        )

        self.operator1_combo.grid(
            row=1,
            column=2,
            sticky="ew",
            padx=5,
        )

        self.value1_entry = ttk.Entry(
            condition1_frame,
            textvariable=self.value1_var,
        )

        self.value1_entry.grid(
            row=1,
            column=3,
            sticky="ew",
            padx=5,
        )

        self.value2_entry = ttk.Entry(
            condition1_frame,
            textvariable=self.value2_var,
            state="disabled",
        )

        self.value2_entry.grid(
            row=1,
            column=4,
            sticky="ew",
            padx=(5, 0),
        )

        # AND / OR
        logical_frame = ttk.LabelFrame(
            conditions,
            text="連接方式",
            padding=10,
        )

        logical_frame.grid(
            row=0,
            column=1,
            sticky="ns",
            padx=6,
        )

        self.logical_combo = ttk.Combobox(
            logical_frame,
            textvariable=self.logical_var,
            values=LOGICAL_OPTIONS,
            state="disabled",
            width=22,
        )

        self.logical_combo.grid(
            row=0,
            column=0,
            sticky="ew",
            pady=(21, 0),
        )

        # 條件 2
        condition2_frame = ttk.LabelFrame(
            conditions,
            text="WHERE 條件 2（選用）",
            padding=10,
        )

        condition2_frame.grid(
            row=0,
            column=2,
            sticky="nsew",
            padx=(6, 0),
        )

        for column in range(5):
            condition2_frame.columnconfigure(
                column,
                weight=1,
            )

        ttk.Label(
            condition2_frame,
            text="欄位",
        ).grid(
            row=0,
            column=0,
            sticky="w",
        )

        ttk.Label(
            condition2_frame,
            text="NOT",
        ).grid(
            row=0,
            column=1,
            sticky="w",
        )

        ttk.Label(
            condition2_frame,
            text="比較方式",
        ).grid(
            row=0,
            column=2,
            sticky="w",
        )

        ttk.Label(
            condition2_frame,
            text="搜尋值",
        ).grid(
            row=0,
            column=3,
            sticky="w",
        )

        ttk.Label(
            condition2_frame,
            text="BETWEEN 第二值",
        ).grid(
            row=0,
            column=4,
            sticky="w",
        )

        self.column2_combo = ttk.Combobox(
            condition2_frame,
            textvariable=self.column2_var,
            state="disabled",
            width=24,
        )

        self.column2_combo.grid(
            row=1,
            column=0,
            sticky="ew",
            padx=(0, 5),
        )

        self.not2_combo = ttk.Combobox(
            condition2_frame,
            textvariable=self.not2_var,
            values=NOT_OPTIONS,
            state="disabled",
            width=15,
        )

        self.not2_combo.grid(
            row=1,
            column=1,
            sticky="ew",
            padx=5,
        )

        self.operator2_combo = ttk.Combobox(
            condition2_frame,
            textvariable=self.operator2_var,
            values=list(OPERATOR_LABELS.values()),
            state="disabled",
            width=22,
        )

        self.operator2_combo.grid(
            row=1,
            column=2,
            sticky="ew",
            padx=5,
        )

        self.value3_entry = ttk.Entry(
            condition2_frame,
            textvariable=self.value3_var,
            state="disabled",
        )

        self.value3_entry.grid(
            row=1,
            column=3,
            sticky="ew",
            padx=5,
        )

        self.value4_entry = ttk.Entry(
            condition2_frame,
            textvariable=self.value4_var,
            state="disabled",
        )

        self.value4_entry.grid(
            row=1,
            column=4,
            sticky="ew",
            padx=(5, 0),
        )

        # -------------------------
        # 排序與按鈕
        # -------------------------

        action_frame = ttk.LabelFrame(
            main,
            text="ORDER BY 與執行",
            padding=10,
        )

        action_frame.grid(
            row=4,
            column=0,
            sticky="ew",
            pady=(0, 8),
        )

        action_frame.columnconfigure(1, weight=1)

        ttk.Label(
            action_frame,
            text="排序欄位",
        ).grid(
            row=0,
            column=0,
            sticky="w",
        )

        self.sort_column_combo = ttk.Combobox(
            action_frame,
            textvariable=self.sort_column_var,
            state="disabled",
            width=26,
        )

        self.sort_column_combo.grid(
            row=1,
            column=0,
            sticky="ew",
            padx=(0, 6),
        )

        ttk.Label(
            action_frame,
            text="方向",
        ).grid(
            row=0,
            column=1,
            sticky="w",
        )

        self.sort_direction_combo = ttk.Combobox(
            action_frame,
            textvariable=self.sort_direction_var,
            values=SORT_DIRECTIONS,
            state="disabled",
            width=16,
        )

        self.sort_direction_combo.grid(
            row=1,
            column=1,
            sticky="w",
            padx=6,
        )

        ttk.Label(
            action_frame,
            text="排序方式",
        ).grid(
            row=0,
            column=2,
            sticky="w",
        )

        self.sort_mode_combo = ttk.Combobox(
            action_frame,
            textvariable=self.sort_mode_var,
            values=SORT_MODES,
            state="disabled",
            width=18,
        )

        self.sort_mode_combo.grid(
            row=1,
            column=2,
            sticky="w",
            padx=6,
        )

        self.run_button = ttk.Button(
            action_frame,
            text="執行查詢",
            style="Primary.TButton",
            command=self.execute_query,
            state="disabled",
        )

        self.run_button.grid(
            row=1,
            column=3,
            padx=(18, 4),
        )

        self.show_all_button = ttk.Button(
            action_frame,
            text="顯示全部",
            command=self.show_all_rows,
            state="disabled",
        )

        self.show_all_button.grid(
            row=1,
            column=4,
            padx=4,
        )

        self.export_button = ttk.Button(
            action_frame,
            text="匯出目前結果 Excel",
            command=self.export_current_result,
            state="disabled",
        )

        self.export_button.grid(
            row=1,
            column=5,
            padx=(4, 0),
        )

        # -------------------------
        # SQL 預覽
        # -------------------------

        sql_frame = ttk.LabelFrame(
            main,
            text="目前產生的 SQL",
            padding=8,
        )

        sql_frame.grid(
            row=5,
            column=0,
            sticky="ew",
            pady=(0, 8),
        )

        sql_frame.columnconfigure(0, weight=1)

        self.sql_text = tk.Text(
            sql_frame,
            height=6,
            background="#111827",
            foreground="#E5E7EB",
            insertbackground="#FFFFFF",
            font=("Consolas", 10),
            wrap="none",
            relief="flat",
            padx=8,
            pady=8,
        )

        self.sql_text.grid(
            row=0,
            column=0,
            sticky="ew",
        )

        self._set_sql_preview(
            "SELECT *\nFROM demo.資料表;"
        )

        # -------------------------
        # 查詢結果
        # -------------------------

        result_frame = ttk.LabelFrame(
            main,
            text="查詢結果",
            padding=6,
        )

        result_frame.grid(
            row=6,
            column=0,
            sticky="nsew",
        )

        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(
            result_frame,
            show="headings",
        )

        self.tree.grid(
            row=0,
            column=0,
            sticky="nsew",
        )

        vertical_scroll = ttk.Scrollbar(
            result_frame,
            orient="vertical",
            command=self.tree.yview,
        )

        vertical_scroll.grid(
            row=0,
            column=1,
            sticky="ns",
        )

        horizontal_scroll = ttk.Scrollbar(
            result_frame,
            orient="horizontal",
            command=self.tree.xview,
        )

        horizontal_scroll.grid(
            row=1,
            column=0,
            sticky="ew",
        )

        self.tree.configure(
            yscrollcommand=vertical_scroll.set,
            xscrollcommand=horizontal_scroll.set,
        )

        self.tree.tag_configure(
            "even",
            background="#F8FAFC",
        )

        self.tree.tag_configure(
            "odd",
            background="#FFFFFF",
        )

        # -------------------------
        # 狀態列
        # -------------------------

        status_frame = ttk.Frame(main)

        status_frame.grid(
            row=7,
            column=0,
            sticky="ew",
            pady=(7, 0),
        )

        status_frame.columnconfigure(0, weight=1)

        ttk.Label(
            status_frame,
            textvariable=self.load_status_var,
        ).grid(
            row=0,
            column=0,
            sticky="w",
        )

        ttk.Label(
            status_frame,
            textvariable=self.result_status_var,
        ).grid(
            row=0,
            column=1,
            sticky="e",
        )

    # --------------------------------------------------------
    # 11-4. 綁定 UI 事件
    # --------------------------------------------------------

    def _bind_events(self) -> None:
        """將下拉選單與輸入框連接到處理函式。"""

        self.table_combo.bind(
            "<<ComboboxSelected>>",
            self.on_table_changed,
        )

        self.column1_combo.bind(
            "<<ComboboxSelected>>",
            self.on_column1_changed,
        )

        self.not1_combo.bind(
            "<<ComboboxSelected>>",
            self.on_preview_changed,
        )

        self.operator1_combo.bind(
            "<<ComboboxSelected>>",
            self.on_operator1_changed,
        )

        self.logical_combo.bind(
            "<<ComboboxSelected>>",
            self.on_preview_changed,
        )

        self.column2_combo.bind(
            "<<ComboboxSelected>>",
            self.on_column2_changed,
        )

        self.not2_combo.bind(
            "<<ComboboxSelected>>",
            self.on_preview_changed,
        )

        self.operator2_combo.bind(
            "<<ComboboxSelected>>",
            self.on_operator2_changed,
        )

        self.sort_column_combo.bind(
            "<<ComboboxSelected>>",
            self.on_preview_changed,
        )

        self.sort_direction_combo.bind(
            "<<ComboboxSelected>>",
            self.on_preview_changed,
        )

        self.sort_mode_combo.bind(
            "<<ComboboxSelected>>",
            self.on_preview_changed,
        )

        # 輸入值時只更新 SQL 預覽；
        # 真正結果仍需按「執行查詢」。
        for entry in [
            self.value1_entry,
            self.value2_entry,
            self.value3_entry,
            self.value4_entry,
        ]:
            entry.bind(
                "<KeyRelease>",
                self.on_preview_changed,
            )

    # --------------------------------------------------------
    # 11-5. UI 選擇值轉換
    # --------------------------------------------------------

    def selected_operator(
        self,
        variable: tk.StringVar,
    ) -> str:
        """將畫面上的中文運算子轉回內部代號。"""

        return LABEL_TO_OPERATOR.get(
            variable.get(),
            "eq",
        )

    def selected_column(
        self,
        display_value: str,
    ) -> str:
        """
        將：
        CustomerID (varchar(10))

        轉回：
        CustomerID
        """

        if display_value in {
            "",
            SECOND_COLUMN_NONE,
            SORT_COLUMN_NONE,
        }:
            return ""

        return self.column_display_map.get(
            display_value,
            display_value,
        )

    def first_uses_not(self) -> bool:
        """條件 1 是否選擇 NOT。"""

        return self.not1_var.get().startswith(
            "NOT"
        )

    def second_uses_not(self) -> bool:
        """條件 2 是否選擇 NOT。"""

        return self.not2_var.get().startswith(
            "NOT"
        )

    def selected_logical_operator(self) -> str:
        """取得 AND 或 OR。"""

        return (
            "OR"
            if self.logical_var.get().startswith("OR")
            else "AND"
        )

    def selected_sort_direction(self) -> str:
        """取得 ASC 或 DESC。"""

        return (
            "DESC"
            if self.sort_direction_var.get().startswith(
                "DESC"
            )
            else "ASC"
        )

    # --------------------------------------------------------
    # 11-6. SQL 預覽
    # --------------------------------------------------------

    def _set_sql_preview(
        self,
        sql_text: str,
    ) -> None:
        """更新唯讀 SQL Text 元件。"""

        self.sql_text.configure(state="normal")
        self.sql_text.delete("1.0", "end")
        self.sql_text.insert("1.0", sql_text)
        self.sql_text.configure(state="disabled")

    def build_sql(self) -> str:
        """
        根據目前 UI 內容組合完整 SQL。

        注意：
        這只負責產生 SQL 文字。
        真正資料由 execute_query() 使用 Python 篩選。
        """

        table_name = self.table_var.get()

        if (
            not table_name
            or table_name not in self.database
        ):
            sql = "SELECT *\nFROM demo.資料表;"
            self._set_sql_preview(sql)
            return sql

        table_data = self.database[table_name]

        raw_condition1 = build_condition_sql(
            self.selected_column(
                self.column1_var.get()
            ),
            self.selected_operator(
                self.operator1_var
            ),
            self.value1_var.get(),
            self.value2_var.get(),
            table_data,
        )

        condition1 = (
            f"NOT ({raw_condition1})"
            if raw_condition1
            and self.first_uses_not()
            else raw_condition1
        )

        raw_condition2 = build_condition_sql(
            self.selected_column(
                self.column2_var.get()
            ),
            self.selected_operator(
                self.operator2_var
            ),
            self.value3_var.get(),
            self.value4_var.get(),
            table_data,
        )

        condition2 = (
            f"NOT ({raw_condition2})"
            if raw_condition2
            and self.second_uses_not()
            else raw_condition2
        )

        sql_lines = [
            "SELECT *",
            f"FROM demo.{table_name}",
        ]

        if condition1 and condition2:
            sql_lines.extend(
                [
                    f"WHERE ({condition1})",
                    (
                        f"  "
                        f"{self.selected_logical_operator()} "
                        f"({condition2})"
                    ),
                ]
            )

        elif condition1:
            sql_lines.append(
                f"WHERE {condition1}"
            )

        elif condition2:
            sql_lines.append(
                f"WHERE {condition2}"
            )

        sort_column = self.selected_column(
            self.sort_column_var.get()
        )

        if sort_column:
            if self.sort_mode_var.get().startswith(
                "轉成數字"
            ):
                sort_expression = (
                    f"CAST({sort_column} "
                    f"AS DECIMAL(30, 10))"
                )
            else:
                sort_expression = sort_column

            sql_lines.append(
                f"ORDER BY "
                f"{sort_expression} "
                f"{self.selected_sort_direction()}"
            )

        sql = "\n".join(sql_lines) + ";"

        self._set_sql_preview(sql)

        return sql

    # --------------------------------------------------------
    # 11-7. 載入 SQL
    # --------------------------------------------------------

    def load_default_sql(self) -> None:
        """啟動時自動載入預設 demo.sql。"""

        default_path = find_default_sql_file()

        if default_path is None:
            self.load_status_var.set(
                "找不到 demo.sql，請按「選擇 SQL 檔」。"
            )
            return

        self.load_sql_file(default_path)

    def choose_sql_file(self) -> None:
        """由使用者選擇其他 .sql 檔案。"""

        selected = filedialog.askopenfilename(
            title="選擇 MySQL Dump 檔案",
            filetypes=[
                ("SQL 檔案", "*.sql"),
                ("所有檔案", "*.*"),
            ],
        )

        if not selected:
            return

        self.load_sql_file(Path(selected))

    def reload_sql_file(self) -> None:
        """重新讀取目前 SQL 檔案。"""

        if (
            self.current_sql_path is None
            or not self.current_sql_path.exists()
        ):
            self.choose_sql_file()
            return

        self.load_sql_file(
            self.current_sql_path
        )

    def load_sql_file(
        self,
        path: Path,
    ) -> None:
        """讀取、解析 SQL，並更新 UI。"""

        self.load_status_var.set(
            f"正在讀取：{path}"
        )

        self.root.update_idletasks()

        try:
            sql_text = path.read_text(
                encoding="utf-8-sig"
            )

            parsed = parse_sql_dump(sql_text)

            if not parsed:
                raise ValueError(
                    "找不到 CREATE TABLE 或 "
                    "INSERT INTO 資料。"
                )

        except Exception as error:
            messagebox.showerror(
                "讀取失敗",
                f"無法讀取 SQL 檔案：\n{error}",
            )

            self.load_status_var.set(
                f"讀取失敗：{error}"
            )
            return

        self.database = parsed
        self.current_sql_path = path

        table_names = self.ordered_table_names()

        self.table_combo.configure(
            values=table_names,
            state="readonly",
        )

        self.table_var.set(table_names[0])

        self.load_status_var.set(
            f"已讀取 {path.name}，"
            f"找到 {len(table_names)} 張資料表。"
        )

        self.populate_columns()
        self.show_all_rows()

    def ordered_table_names(self) -> list[str]:
        """依指定順序排列資料表名稱。"""

        names = list(self.database.keys())

        preferred = [
            name
            for name in PREFERRED_TABLE_ORDER
            if name in names
        ]

        others = sorted(
            name
            for name in names
            if name not in PREFERRED_TABLE_ORDER
        )

        return preferred + others

    # --------------------------------------------------------
    # 11-8. 資料表與欄位選單
    # --------------------------------------------------------

    def on_table_changed(
        self,
        _event: tk.Event | None = None,
    ) -> None:
        """更換資料表時重新產生欄位選單。"""

        self.populate_columns()
        self.show_all_rows()

    def populate_columns(self) -> None:
        """將目前資料表欄位放進三個下拉選單。"""

        table_name = self.table_var.get()

        if table_name not in self.database:
            return

        table_data = self.database[table_name]

        self.column_display_map = {}

        displays: list[str] = []

        for column in table_data.columns:
            type_name = table_data.types.get(
                column,
                "unknown",
            )

            display = (
                f"{column} ({type_name})"
            )

            self.column_display_map[
                display
            ] = column

            displays.append(display)

        self.column1_combo.configure(
            values=displays,
            state="readonly",
        )

        self.column2_combo.configure(
            values=[
                SECOND_COLUMN_NONE,
                *displays,
            ],
            state="readonly",
        )

        self.sort_column_combo.configure(
            values=[
                SORT_COLUMN_NONE,
                *displays,
            ],
            state="readonly",
        )

        if displays:
            self.column1_var.set(displays[0])
        else:
            self.column1_var.set("")

        self.not1_combo.configure(
            state="readonly"
        )

        self.operator1_combo.configure(
            state="readonly"
        )

        self.sort_column_combo.configure(
            state="readonly"
        )

        self.sort_direction_combo.configure(
            state="readonly"
        )

        self.sort_mode_combo.configure(
            state="readonly"
        )

        self.run_button.configure(
            state="normal"
        )

        self.show_all_button.configure(
            state="normal"
        )

        self.export_button.configure(
            state="normal"
        )

        self.reset_query_controls()
        self.update_condition1_controls()
        self.update_condition2_controls()

    def reset_query_controls(self) -> None:
        """切換資料表時清除舊查詢條件。"""

        self.not1_var.set(NOT_OPTIONS[0])
        self.value1_var.set("")
        self.value2_var.set("")

        self.column2_var.set(
            SECOND_COLUMN_NONE
        )
        self.not2_var.set(NOT_OPTIONS[0])
        self.operator2_var.set(
            OPERATOR_LABELS["like"]
        )
        self.value3_var.set("")
        self.value4_var.set("")
        self.logical_var.set(
            LOGICAL_OPTIONS[0]
        )

        self.sort_column_var.set(
            SORT_COLUMN_NONE
        )
        self.sort_direction_var.set(
            SORT_DIRECTIONS[0]
        )
        self.sort_mode_var.set(
            SORT_MODES[0]
        )

        self.set_default_operator_for_column(
            self.column1_var.get(),
            self.operator1_var,
        )

    def set_default_operator_for_column(
        self,
        display_column: str,
        operator_variable: tk.StringVar,
    ) -> None:
        """
        數字欄位預設使用等於，
        文字欄位預設使用 LIKE。
        """

        table_name = self.table_var.get()

        if table_name not in self.database:
            return

        column_name = self.selected_column(
            display_column
        )

        type_name = self.database[
            table_name
        ].types.get(
            column_name,
            "",
        )

        operator_variable.set(
            OPERATOR_LABELS[
                "eq"
                if is_numeric_type(type_name)
                else "like"
            ]
        )

    # --------------------------------------------------------
    # 11-9. 條件元件狀態
    # --------------------------------------------------------

    def on_column1_changed(
        self,
        _event: tk.Event | None = None,
    ) -> None:
        """第一欄位改變時更新預設運算子。"""

        self.set_default_operator_for_column(
            self.column1_var.get(),
            self.operator1_var,
        )

        self.update_condition1_controls()

    def on_operator1_changed(
        self,
        _event: tk.Event | None = None,
    ) -> None:
        """第一運算子改變時更新輸入框。"""

        self.update_condition1_controls()

    def update_condition1_controls(self) -> None:
        """啟用或停用條件 1 的輸入框。"""

        operator = self.selected_operator(
            self.operator1_var
        )

        no_value_needed = operator in {
            "isnull",
            "notnull",
        }

        needs_second_value = (
            operator == "between"
        )

        self.value1_entry.configure(
            state=(
                "disabled"
                if no_value_needed
                else "normal"
            )
        )

        self.value2_entry.configure(
            state=(
                "normal"
                if needs_second_value
                else "disabled"
            )
        )

        self.build_sql()

    def on_column2_changed(
        self,
        _event: tk.Event | None = None,
    ) -> None:
        """第二欄位改變時啟用或停用整組條件。"""

        if (
            self.column2_var.get()
            == SECOND_COLUMN_NONE
        ):
            self.not2_var.set(NOT_OPTIONS[0])
            self.operator2_var.set(
                OPERATOR_LABELS["like"]
            )
            self.value3_var.set("")
            self.value4_var.set("")
        else:
            self.set_default_operator_for_column(
                self.column2_var.get(),
                self.operator2_var,
            )

        self.update_condition2_controls()

    def on_operator2_changed(
        self,
        _event: tk.Event | None = None,
    ) -> None:
        """第二運算子改變時更新輸入框。"""

        self.update_condition2_controls()

    def update_condition2_controls(self) -> None:
        """啟用或停用條件 2 的控制元件。"""

        active = (
            self.column2_var.get()
            != SECOND_COLUMN_NONE
        )

        operator = self.selected_operator(
            self.operator2_var
        )

        no_value_needed = operator in {
            "isnull",
            "notnull",
        }

        needs_second_value = (
            operator == "between"
        )

        combo_state = (
            "readonly"
            if active
            else "disabled"
        )

        self.logical_combo.configure(
            state=combo_state
        )

        self.not2_combo.configure(
            state=combo_state
        )

        self.operator2_combo.configure(
            state=combo_state
        )

        self.value3_entry.configure(
            state=(
                "normal"
                if active
                and not no_value_needed
                else "disabled"
            )
        )

        self.value4_entry.configure(
            state=(
                "normal"
                if active
                and needs_second_value
                else "disabled"
            )
        )

        self.build_sql()

    def on_preview_changed(
        self,
        _event: tk.Event | None = None,
    ) -> None:
        """任一條件改變時即時更新 SQL 預覽。"""

        self.build_sql()

    # --------------------------------------------------------
    # 11-10. 執行查詢
    # --------------------------------------------------------

    def execute_query(self) -> None:
        """
        依目前條件篩選與排序，
        然後將結果放進 Treeview。
        """

        table_name = self.table_var.get()

        if table_name not in self.database:
            return

        table_data = self.database[table_name]

        column1 = self.selected_column(
            self.column1_var.get()
        )

        column2 = self.selected_column(
            self.column2_var.get()
        )

        operator1 = self.selected_operator(
            self.operator1_var
        )

        operator2 = self.selected_operator(
            self.operator2_var
        )

        logical_operator = (
            self.selected_logical_operator()
        )

        filtered: list[list[Any]] = []

        for row in table_data.rows:
            result1 = evaluate_condition(
                row,
                table_data,
                column1,
                operator1,
                self.value1_var.get(),
                self.value2_var.get(),
            )

            result2 = evaluate_condition(
                row,
                table_data,
                column2,
                operator2,
                self.value3_var.get(),
                self.value4_var.get(),
            )

            result1 = apply_sql_not(
                result1,
                self.first_uses_not(),
            )

            result2 = apply_sql_not(
                result2,
                self.second_uses_not(),
            )

            # 兩個條件都未啟用。
            if (
                result1 is INACTIVE_CONDITION
                and result2 is INACTIVE_CONDITION
            ):
                final_result: bool | None = True

            # 只有條件 1。
            elif result2 is INACTIVE_CONDITION:
                final_result = result1  # type: ignore[assignment]

            # 只有條件 2。
            elif result1 is INACTIVE_CONDITION:
                final_result = result2  # type: ignore[assignment]

            # 兩個條件都有使用。
            else:
                final_result = combine_sql_results(
                    result1,  # type: ignore[arg-type]
                    result2,  # type: ignore[arg-type]
                    logical_operator,
                )

            # SQL WHERE 只保留最終結果為 TRUE 的 Row。
            if final_result is True:
                filtered.append(row)

        sort_column = self.selected_column(
            self.sort_column_var.get()
        )

        self.current_result = sort_rows(
            filtered,
            table_data,
            sort_column,
            self.selected_sort_direction(),
            self.sort_mode_var.get().startswith(
                "轉成數字"
            ),
        )

        self.last_executed_sql = self.build_sql()

        self.render_result(
            table_name,
            table_data,
            self.current_result,
        )

    def show_all_rows(self) -> None:
        """清除條件並顯示目前資料表全部資料。"""

        table_name = self.table_var.get()

        if table_name not in self.database:
            return

        table_data = self.database[table_name]

        self.not1_var.set(NOT_OPTIONS[0])
        self.value1_var.set("")
        self.value2_var.set("")

        self.column2_var.set(
            SECOND_COLUMN_NONE
        )
        self.not2_var.set(NOT_OPTIONS[0])
        self.operator2_var.set(
            OPERATOR_LABELS["like"]
        )
        self.value3_var.set("")
        self.value4_var.set("")

        self.logical_var.set(
            LOGICAL_OPTIONS[0]
        )

        self.sort_column_var.set(
            SORT_COLUMN_NONE
        )
        self.sort_direction_var.set(
            SORT_DIRECTIONS[0]
        )
        self.sort_mode_var.set(
            SORT_MODES[0]
        )

        self.update_condition1_controls()
        self.update_condition2_controls()

        self.current_result = list(
            table_data.rows
        )

        self.last_executed_sql = self.build_sql()

        self.render_result(
            table_name,
            table_data,
            self.current_result,
        )

    # --------------------------------------------------------
    # 11-11. 顯示結果
    # --------------------------------------------------------

    def render_result(
        self,
        table_name: str,
        table_data: TableData,
        rows: list[list[Any]],
    ) -> None:
        """將查詢結果顯示在 Treeview。"""

        # 清除舊資料。
        for item in self.tree.get_children():
            self.tree.delete(item)

        columns = table_data.columns

        self.tree.configure(
            columns=columns,
            displaycolumns=columns,
        )

        for column_index, column in enumerate(
            columns
        ):
            self.tree.heading(
                column,
                text=column,
            )

            # 依欄位與部分資料估算寬度。
            sample_values = [
                row[column_index]
                for row in rows[:150]
                if column_index < len(row)
            ]

            maximum_length = len(column)

            for value in sample_values:
                value_text = (
                    "NULL"
                    if value is None
                    else str(value)
                )

                maximum_length = max(
                    maximum_length,
                    len(value_text),
                )

            width = min(
                max(maximum_length * 8 + 24, 90),
                360,
            )

            self.tree.column(
                column,
                width=width,
                minwidth=70,
                stretch=True,
                anchor="w",
            )

        for row_index, row in enumerate(rows):
            display_values = [
                (
                    "NULL"
                    if value is None
                    else value
                )
                for value in row
            ]

            self.tree.insert(
                "",
                "end",
                values=display_values,
                tags=(
                    "even"
                    if row_index % 2 == 0
                    else "odd",
                ),
            )

        self.result_status_var.set(
            f"資料表：{table_name}　"
            f"結果：{len(rows)} / "
            f"{len(table_data.rows)} 筆"
        )

    # --------------------------------------------------------
    # 11-12. Excel 匯出
    # --------------------------------------------------------

    def export_current_result(self) -> None:
        """匯出目前畫面中已執行完成的查詢結果。"""

        table_name = self.table_var.get()

        if table_name not in self.database:
            messagebox.showwarning(
                "無法匯出",
                "目前沒有可匯出的資料表。",
            )
            return

        output_name = (
            f"{table_name}_查詢結果.xlsx"
        )

        selected = filedialog.asksaveasfilename(
            title="儲存 Excel",
            defaultextension=".xlsx",
            initialfile=output_name,
            filetypes=[
                ("Excel 活頁簿", "*.xlsx"),
            ],
        )

        if not selected:
            return

        try:
            export_result_to_excel(
                Path(selected),
                table_name,
                self.database[table_name].columns,
                self.current_result,
                self.last_executed_sql,
            )
        except Exception as error:
            messagebox.showerror(
                "匯出失敗",
                f"無法建立 Excel：\n{error}",
            )
            return

        self.load_status_var.set(
            f"已匯出 Excel：{selected}"
        )

        messagebox.showinfo(
            "匯出完成",
            f"Excel 已儲存：\n{selected}",
        )


# ============================================================
# 12. 程式進入點
# ============================================================

def main() -> None:
    """建立 tkinter 視窗並啟動事件迴圈。"""

    root = tk.Tk()
    DemoSqlQueryApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
